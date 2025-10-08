import openpyxl
import requests
from supabase import create_client, Client
import json
import re
import subprocess
import pandas as pd
from geopy.geocoders import GoogleV3

# Supabase credentials
SUPABASE_URL = "https://xrbfefmwiymnnpujtcbg.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InhyYmZlZm13aXltbm5wdWp0Y2JnIiwicm9sZSI6ImFub24iLCJpYXQiOjE2NzU5OTAzNTksImV4cCI6MTk5MTU2NjM1OX0.E1HZXgTpL9bb38Oc3CAwBOLXtGSK1t9ZbgtRcWnqRUs"
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# Google API Key
api_key = "AIzaSyDUZVelPmEmOGYeOWRaiy4dVCt-gLnEgWY"

# Function to get Google Plus Code
def get_plus_code(latitude, longitude):
    plus_code_url = f'https://plus.codes/api?address={latitude},{longitude}'
    plus_code_response = requests.get(plus_code_url)
    plus_code_data = plus_code_response.json()
    
    if 'plus_code' in plus_code_data:
        return plus_code_data['plus_code']['global_code']
    else:
        return "Plus Code not found for the given coordinates"
    
test_plus_code = get_plus_code(42.8939674, -85.4898443)
print(test_plus_code)

# Function to get geocode data
def get_geocode_data(address):
    geocode_url = f'https://maps.googleapis.com/maps/api/geocode/json?address={address}&key={api_key}'
    geocode_response = requests.get(geocode_url)
    geocode_data = geocode_response.json()
    
    if geocode_data['status'] == 'OK':
        location = geocode_data['results'][0]['geometry']['location']
        latitude = location['lat']
        longitude = location['lng']
        return latitude, longitude
    else:
        return None, None

# Function to process the name for intersections
def process_name(address, name):
    if name:
        return name
    else:
        # Define the list of common street suffixes
        street_suffixes = ["Dr", "St", "Ct", "Ave", "Blvd", "Ln", "Rd", "Terr", "Pl", "Way"]
        pattern = r'^(\d+)\s+([\w\s]+?)\s+(' + '|'.join(street_suffixes) + ')'
        match = re.search(pattern, address)
        if match:
            house_number = match.group(1)
            street_name = match.group(2)
            return f"{house_number} {street_name}"
        
        # Special case: Check for address ending with comma
        comma_pattern = r'^(\d+)\s+(.*?),.*$'
        match = re.search(comma_pattern, address)
        if match:
            house_number = match.group(1)
            street_name = match.group(2)
            return f"{house_number} {street_name}"
        
        return address

# Function to format intersection addresses
def format_intersection(address):
    if "&" in address:
        parts = address.split("&")
        if len(parts) == 2:
            return f"{parts[0].strip()} and {parts[1].strip()}"
    return address

# Function to parse the stdout of the subprocess
def parse_stdout(output):
    lines = output.split('\n')
    results = []
    for line in lines:
        if line.startswith("Address: "):
            try:
                address_match = re.search(r"Address: (.*?), OSM Street Name:", line)
                osm_street_name_match = re.search(r"OSM Street Name: (.*?), Maxspeed:", line)
                maxspeed_match = re.search(r"Maxspeed: (.*?), Road Type:", line)
                road_type_match = re.search(r"Road Type: (.*?), Difference:", line)
                difference_match = re.search(r"Difference: (.*)", line)

                if address_match and osm_street_name_match and maxspeed_match and road_type_match and difference_match:
                    address = address_match.group(1)
                    osm_street_name = osm_street_name_match.group(1)
                    maxspeed = maxspeed_match.group(1)
                    road_type = road_type_match.group(1)
                    difference = difference_match.group(1)
                    results.append((address, osm_street_name, maxspeed, road_type, difference))
                else:
                    print(f"Error parsing line: {line}. Some parts of the line did not match the expected pattern.")
            except (IndexError, ValueError) as e:
                print(f"Error parsing line: {line}. Error: {e}")
    return results

# Function to extract coordinates from Plus Code using the Google Plus Codes API
def get_coordinates_from_plus_code(plus_code, api_key):
    try:
        geolocator = GoogleV3(api_key)
        location = geolocator.geocode(plus_code)
        if location:
            return location.latitude, location.longitude
        else:
            print("Location not found for Plus Code.")
            return None, None
    except Exception as e:
        print(f"Error decoding Plus Code: {e}")
        return None, None

# Load Stops.xlsx
stops_wb = openpyxl.load_workbook('Stops.xlsx')
stops_ws = stops_wb.active

# Load all_stops_speed_limit_road_type.xlsx
all_addresses_wb = openpyxl.load_workbook('all_stops_speedLimit_roadType.xlsx')
all_addresses_ws = all_addresses_wb.active

# Create a dictionary for all_addresses lookup
all_addresses_dict = {}
for row in all_addresses_ws.iter_rows(min_row=2, values_only=True):
    address, _, maxspeed, road_type, *_ = row  # Unpack only relevant columns
    all_addresses_dict[address] = {
        "Maxspeed": maxspeed,
        "Road Type": road_type
    }

# Fetch existing addresses from the stops table
existing_stops_response = supabase.table('stops').select('address').execute()
existing_addresses = {record['address'] for record in existing_stops_response.data}
print(f"existing address: {existing_addresses}")

# Fetch all IDs greater than 50
existing_ids_response = supabase.table('stops').select('id').gt('id', 50).execute()
existing_ids = {record['id'] for record in existing_ids_response.data}

# Function to find the next available ID above 50
def get_next_id(existing_ids, start_id=51):
    current_id = start_id
    while current_id in existing_ids:
        current_id += 1
    return current_id

new_addresses = []

# Process Stops.xlsx and insert data into Supabase
for row in stops_ws.iter_rows(min_row=2, values_only=True):
    address, name, plus_code, location, *_ = row  # Ignore columns E and F
    print(plus_code)

    # Check if the address already exists in the stops table
    if address in existing_addresses:
        continue
    else:
        print(f"new address: {address}")

    # Process the name for intersections
    name = process_name(address, name)

    # Format the address if it's an intersection
    formatted_address = format_intersection(address)
    print(address)

    # Extract coordinates from the detailed Plus Code
    try:
        latitude, longitude = get_coordinates_from_plus_code(plus_code, api_key)
        print(latitude)
        print(longitude)
    except requests.exceptions.RequestException as e:
        print(f"Request error for address: {address}. Error: {e}, skipping entry.")
        continue

    if latitude and longitude:
        geo_location = f"POINT({longitude} {latitude})"
    else:
        print(f"Failed to extract coordinates from Plus Code for address: {address}, skipping entry.")
        continue

    # Get the Plus Code using latitude and longitude
    try:
        google_plus_code = get_plus_code(latitude, longitude)
    except requests.exceptions.RequestException as e:
        print(f"Request error for Plus Code: {address}. Error: {e}, skipping entry.")
        continue

    # Get metadata from all_stops_speed_limit_road_type.xlsx
    all_addresses_metadata = all_addresses_dict.get(address, None)
    if all_addresses_metadata:
        maxspeed = all_addresses_metadata['Maxspeed']
        road_type = all_addresses_metadata['Road Type']
    else:
        # Add new address to the list to fetch speed limit and road type
        new_addresses.append(address)
        print(new_addresses)
        maxspeed = 'Unknown'
        road_type = 'Unknown'

    # Update the metadata to include the location
    metadata = [
        {
            "Speed Limit": maxspeed,
            "Road Type": road_type,
            "Location": location  # Directly use the location from the spreadsheet
        }
    ]

    # Get the next available ID
    current_id = get_next_id(existing_ids)
    existing_ids.add(current_id)  # Add the new ID to the set of existing IDs

    # Insert data into Supabase
    data = {
        "id": current_id,  # Use the current_id for insertion
        "address": address,
        "name": name,
        "google_plus_code": google_plus_code,  # Insert Google Plus Code found by the function
        "detailed_plus_code": plus_code,  # Insert Plus Code from the spreadsheet
        "geo_location": geo_location,
        "metadata": json.dumps(metadata),  # Ensure the metadata is JSON serialized
    }

    try:
        response = supabase.table('stops').insert(data).execute()
        print(f"Inserted data: {data}")
    except Exception as e:
        print(f"Error inserting data: {data}. Error: {e}")
        continue

# Fetch speed limits and road types for new addresses
if new_addresses:
    with open('addresses.txt', 'w') as f:
        for address in new_addresses:
            f.write(f"{address}\n")

    result = subprocess.run(
        ['python', 'speedLimitRoadType.py'],
        capture_output=True,
        text=True,
    )
    print(f"result.stdout: {result.stdout}")
    print(f"result.stderr: {result.stderr}")

    # Parse the output
    parsed_results = parse_stdout(result.stdout)
    print(f"parsed_results: {parsed_results}")

    for result in parsed_results:
        address, osm_street_name, maxspeed, road_type, difference = result

        # Find the corresponding stop in the Supabase table and update metadata
        try:
            stop_data = supabase.table('stops').select('id', 'metadata').eq('address', address).execute().data[0]
            stop_id = stop_data['id']
            metadata = json.loads(stop_data['metadata'])
            metadata[0]['Speed Limit'] = maxspeed
            metadata[0]['Road Type'] = road_type

            supabase.table('stops').update({"metadata": json.dumps(metadata)}).eq('id', stop_id).execute()
        except Exception as e:
            print(f"Error updating metadata for address: {address}. Error: {e}")
